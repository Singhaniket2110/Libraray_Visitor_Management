from flask import Blueprint, jsonify, request, render_template, make_response,session
from functools import wraps
from datetime import datetime, timedelta
import json
import io
import csv
import pandas as pd
import jwt

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# ✅ USE DIRECT API INSTEAD OF SUPABASE_DB
from backend.supabase_direct import SupabaseDirect as Database

# Import models
from backend.models.visitor_model import get_all_visitors, get_today_visitors, get_filtered_visitors, get_visitors_by_date_range
from backend.config import Config

# ==================== JWT HELPER FUNCTIONS ====================

def create_jwt_token(username):
    """Create JWT token for authentication"""
    payload = {
        'username': username,
        'exp': datetime.utcnow() + timedelta(hours=12),
        'iat': datetime.utcnow()
    }
    token = jwt.encode(payload, Config.JWT_SECRET_KEY, algorithm='HS256')
    return token

def verify_jwt_token(token):
    """Verify JWT token"""
    try:
        payload = jwt.decode(token, Config.JWT_SECRET_KEY, algorithms=['HS256'])
        return payload['username']
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

# ==================== LOGIN REQUIRED DECORATOR ====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Get token from cookie
        token = request.cookies.get('admin_token')
        
        if not token:
            return jsonify({'error': 'Authentication required', 'redirect': '/admin/login'}), 401
        
        # Verify token
        username = verify_jwt_token(token)
        if not username:
            return jsonify({'error': 'Invalid or expired token', 'redirect': '/admin/login'}), 401
        
        # Add username to request context
        request.admin_username = username
        return f(*args, **kwargs)
    
    return decorated_function

# ==================== AUTHENTICATION ROUTES ====================

@admin_bp.route('/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page and authentication"""
    if request.method == 'GET':
        return render_template('login.html')
    
    # POST: Process login
    try:
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not username or not password:
            return jsonify({
                'success': False,
                'error': 'Username and password are required'
            }), 400
        
        # ✅ USE DIRECT API
        admin = Database.admin_login(username, password)
        
        if admin:
            # Create JWT token
            token = create_jwt_token(username)
            
            # Create response with token in cookie
            response = make_response(jsonify({
                'success': True,
                'message': 'Login successful',
                'redirect': '/admin/dashboard'
            }))
            
            # Set token in HTTP-only cookie
            response.set_cookie(
                'admin_token',
                token,
                httponly=True,
                secure=True,
                samesite='Lax',
                max_age=43200  # 12 hours
            )
            
            return response, 200
        else:
            return jsonify({
                'success': False,
                'error': 'Invalid username or password'
            }), 401
            
    except Exception as e:
        print(f"❌ Login error: {e}")
        return jsonify({'success': False, 'error': 'Login failed. Please try again.'}), 500

@admin_bp.route('/logout')
def admin_logout():
    """Logout admin"""
    try:
        # Clear Flask session
        session.clear()
        
        # Create response
        response = make_response(jsonify({
            'success': True, 
            'message': 'Logged out successfully',
            'redirect': '/'
        }))
        
        # Clear all cookies
        response.set_cookie('admin_token', '', expires=0)
        response.set_cookie('session', '', expires=0)
        response.set_cookie('library_session', '', expires=0)
        
        print("✅ Admin session cleared, redirecting to home")
        return response
        
    except Exception as e:
        print(f"❌ Logout error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
@admin_bp.route('/check_session')
def check_session():
    """Check if admin is logged in"""
    token = request.cookies.get('admin_token')
    
    if not token:
        return jsonify({'logged_in': False, 'username': None})
    
    username = verify_jwt_token(token)
    
    if username:
        return jsonify({'logged_in': True, 'username': username})
    else:
        return jsonify({'logged_in': False, 'username': None})

# ==================== DASHBOARD ROUTES ====================

@admin_bp.route('/dashboard')
@login_required
def admin_dashboard():
    """Admin dashboard"""
    return render_template('dashboard.html')

# ==================== ADD VISITOR (ADMIN ONLY) ====================

@admin_bp.route('/add_visitor', methods=['POST'])
@login_required
def add_visitor_admin():
    """Add visitor with custom date and time - Admin only"""
    try:
        data = request.json
        
        # Validate required fields
        required_fields = ['name', 'roll_no', 'level', 'purpose', 'visit_date', 'entry_time']
        for field in required_fields:
            if not data.get(field):
                return jsonify({"error": f"{field} is required"}), 400
        
        # Additional validation for JC
        if data.get('level') == 'JC':
            if not data.get('jc_year') or not data.get('jc_stream'):
                return jsonify({"error": "JC Year and Stream are required for JC students"}), 400
        else:
            if not data.get('course'):
                return jsonify({"error": "Course is required for UG/PG students"}), 400
        
        # Calculate visit day from date
        visit_date_obj = datetime.strptime(data['visit_date'], '%Y-%m-%d')
        visit_day = visit_date_obj.strftime('%A')
        
        # Prepare data for direct API
        visitor_data = {
            'name': data['name'].strip(),
            'roll_no': data['roll_no'].strip().upper(),
            'level': data['level'],
            'purpose': data['purpose'],
            'visit_date': data['visit_date'],
            'entry_time': data['entry_time'],
            'visit_day': visit_day
        }
        
        # Add optional fields
        if data.get('exit_time'):
            visitor_data['exit_time'] = data['exit_time']
        
        # Add course and year/stream
        if data['level'] == 'JC':
            visitor_data['course'] = data.get('course', 'Junior College')
            visitor_data['jc_year'] = data.get('jc_year')
            visitor_data['jc_stream'] = data.get('jc_stream')
        else:
            visitor_data['course'] = data['course']
            if data.get('year'):
                visitor_data['year'] = data.get('year')
        
        # ✅ USE DIRECT API
        result = Database.insert_visitor(visitor_data)
        
        if result:
            return jsonify({
                "success": True,
                "message": "Visitor added successfully",
                "visitor_id": result.get('id')
            }), 201
        else:
            return jsonify({"error": "Failed to add visitor"}), 500
        
    except Exception as e:
        print(f"Error adding visitor: {e}")
        return jsonify({"error": f"Failed to add visitor: {str(e)}"}), 500

# ==================== VISITOR MANAGEMENT ROUTES ====================

@admin_bp.route('/visitors/today')
@login_required
def today_visitors():
    """Get today's visitors"""
    try:
        # ✅ USE DIRECT API
        visitors = Database.get_today_visitors()
        return jsonify(visitors), 200
    except Exception as e:
        print(f"Error getting today's visitors: {e}")
        return jsonify({"error": "Error fetching data"}), 500

@admin_bp.route('/visitors/all')
@login_required
def all_visitors():
    """Get all visitors"""
    try:
        # ✅ USE DIRECT API
        visitors = Database.get_all_visitors()
        return jsonify(visitors), 200
    except Exception as e:
        print(f"Error getting all visitors: {e}")
        return jsonify({"error": "Error fetching data"}), 500

@admin_bp.route('/visitors/filter')
@login_required
def filtered_visitors():
    """Get filtered visitors"""
    try:
        level = request.args.get('level', '')
        date = request.args.get('date', '')
        
        # For now, get all and filter locally
        visitors = Database.get_all_visitors()
        
        # Apply filters
        if level:
            visitors = [v for v in visitors if v.get('level') == level]
        if date:
            visitors = [v for v in visitors if str(v.get('visit_date')) == date]
        
        return jsonify(visitors), 200
    except Exception as e:
        print(f"Error filtering visitors: {e}")
        return jsonify({"error": "Error filtering data"}), 500

# ==================== ANALYTICS ROUTES ====================

@admin_bp.route('/analytics/advanced')
@login_required
def advanced_analytics():
    """Get advanced analytics data"""
    try:
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # Validate dates
        if not start_date or start_date == 'null':
            start_date = datetime.now().date().isoformat()
        if not end_date or end_date == 'null':
            end_date = datetime.now().date().isoformat()
        
        # ✅ USE DIRECT API
        visitors = Database.get_visitors_by_date_range(start_date, end_date)
        
        # Basic stats
        from collections import Counter
        
        total = len(visitors)
        active = sum(1 for v in visitors if v.get('exit_time') is None)
        
        # Level distribution
        level_counts = Counter(v.get('level', 'Unknown') for v in visitors)
        
        # Course distribution
        course_counts = Counter()
        for v in visitors:
            if v.get('level') == 'JC':
                course = v.get('jc_stream', 'Unknown')
            else:
                course = v.get('course', 'Unknown')
            course_counts[course] += 1
        
        # Purpose distribution
        purpose_counts = Counter(v.get('purpose', 'Other') for v in visitors)
        
        # Daily trend
        daily_counts = {}
        for v in visitors:
            date = v.get('visit_date')
            if date:
                date_str = str(date)
                daily_counts[date_str] = daily_counts.get(date_str, 0) + 1
        
        sorted_dates = sorted(daily_counts.keys())
        
        # Peak hours analysis
        hour_counts = {hour: 0 for hour in range(8, 21)}
        for v in visitors:
            if v.get('entry_time'):
                try:
                    entry_time = str(v['entry_time'])
                    hour = int(entry_time.split(':')[0])
                    if 8 <= hour <= 20:
                        hour_counts[hour] = hour_counts.get(hour, 0) + 1
                except:
                    pass
        
        # Average duration
        durations = []
        for v in visitors:
            if v.get('entry_time') and v.get('exit_time'):
                try:
                    entry_str = str(v['entry_time'])
                    exit_str = str(v['exit_time'])
                    
                    entry = datetime.strptime(entry_str, '%H:%M:%S')
                    exit_time = datetime.strptime(exit_str, '%H:%M:%S')
                    
                    if exit_time < entry:
                        exit_time = datetime.combine(
                            exit_time.date() + timedelta(days=1),
                            exit_time.time()
                        )
                    
                    duration = (exit_time - entry).total_seconds() / 60
                    if duration >= 0:
                        durations.append(duration)
                except:
                    pass
        
        avg_duration = sum(durations) / len(durations) if durations else 0
        
        response_data = {
            'stats': {
                'total': total,
                'active': active,
                'avgDuration': round(avg_duration, 1)
            },
            'levelData': {
                'labels': list(level_counts.keys()),
                'values': list(level_counts.values()),
                'colors': ['#f59e0b', '#10b981', '#8b5cf6']
            },
            'courseData': {
                'labels': [c[0] for c in course_counts.most_common(10)],
                'values': [c[1] for c in course_counts.most_common(10)]
            },
            'purposeData': {
                'labels': list(purpose_counts.keys()),
                'values': list(purpose_counts.values())
            },
            'dailyTrend': {
                'labels': [str(d) for d in sorted_dates],
                'values': [daily_counts[d] for d in sorted_dates]
            },
            'peakHours': {
                'labels': [f"{h}:00" for h in range(8, 21)],
                'values': [hour_counts[h] for h in range(8, 21)]
            },
            'visitors': visitors
        }
        
        return jsonify(response_data), 200
        
    except Exception as e:
        print(f"Analytics error: {e}")
        return jsonify({"error": f"Error fetching analytics: {str(e)}"}), 500

# ==================== DATA IMPORT/EXPORT ROUTES ====================

@admin_bp.route('/import_data', methods=['POST'])
@login_required
def import_data():
    """Import data from CSV or Excel"""
    try:
        from flask import send_file
        
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Check file extension
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        elif file.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file)
        else:
            return jsonify({"error": "Unsupported file format"}), 400
        
        # Validate required columns
        required_columns = ['name', 'roll_no', 'level', 'purpose']
        for col in required_columns:
            if col not in df.columns:
                return jsonify({"error": f"Missing required column: {col}"}), 400
        
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                visitor_data = {
                    'name': str(row['name']).strip(),
                    'roll_no': str(row['roll_no']).strip().upper(),
                    'level': str(row['level']).strip().upper(),
                    'purpose': str(row['purpose']).strip(),
                    'course': str(row['course']).strip() if 'course' in df.columns and pd.notna(row.get('course')) else 'Not Specified'
                }
                
                # Add year/stream based on level
                if visitor_data['level'] == 'JC':
                    if 'jc_year' in df.columns and pd.notna(row.get('jc_year')):
                        visitor_data['jc_year'] = str(row['jc_year']).strip()
                    if 'jc_stream' in df.columns and pd.notna(row.get('jc_stream')):
                        visitor_data['jc_stream'] = str(row['jc_stream']).strip()
                elif 'year' in df.columns and pd.notna(row.get('year')):
                    visitor_data['year'] = str(row['year']).strip()
                
                # ✅ USE DIRECT API
                result = Database.insert_visitor(visitor_data)
                
                if result:
                    imported_count += 1
                else:
                    errors.append(f"Row {index + 1}: Failed to insert")
                
            except Exception as e:
                errors.append(f"Row {index + 1}: {str(e)}")
        
        return jsonify({
            "success": True,
            "message": f"Imported {imported_count} records successfully",
            "errors": errors[:10] if errors else []
        }), 200
        
    except Exception as e:
        print(f"Import error: {e}")
        return jsonify({"error": f"Import failed: {str(e)}"}), 500

@admin_bp.route('/export_data', methods=['GET'])
@login_required
def export_data():
    """Export visitor data — supports month, custom range, or all-time. CSV or Excel."""
    try:
        from flask import send_file
        from calendar import monthrange

        format_type = request.args.get('format', 'csv')   # csv | excel
        scope       = request.args.get('scope', 'range')  # range | month | year | all
        start_date  = request.args.get('start_date', '')
        end_date    = request.args.get('end_date', '')
        month       = request.args.get('month', '')        # YYYY-MM
        year        = request.args.get('year', '')         # YYYY

        # ── resolve date range from scope ─────────────────────────
        label = 'all'
        if scope == 'month' and month:
            y, m = int(month.split('-')[0]), int(month.split('-')[1])
            start_date = f"{y}-{m:02d}-01"
            end_date   = f"{y}-{m:02d}-{monthrange(y, m)[1]:02d}"
            label = month  # e.g. "2025-03"
        elif scope == 'year' and year:
            start_date = f"{year}-01-01"
            end_date   = f"{year}-12-31"
            label = year
        elif scope == 'range' and start_date and end_date:
            label = f"{start_date}_to_{end_date}"
        # else scope == 'all' → fetch everything

        if start_date and end_date:
            visitors = Database.get_visitors_by_date_range(start_date, end_date)
        else:
            visitors = Database.get_all_visitors()

        # ── shared row builder ─────────────────────────────────────
        HEADERS = ['ID','Name','Roll No','Level','Course','Year',
                   'JC Year','JC Stream','Purpose',
                   'Entry Time','Exit Time','Visit Date','Day','Duration (min)']

        def build_row(v):
            # calculate duration
            duration = ''
            if v.get('entry_time') and v.get('exit_time'):
                try:
                    from datetime import datetime as dt
                    e = dt.strptime(str(v['entry_time']), '%H:%M:%S')
                    x = dt.strptime(str(v['exit_time']),  '%H:%M:%S')
                    duration = str(round((x - e).total_seconds() / 60))
                except Exception:
                    pass
            return [
                v.get('id',''),        v.get('name',''),
                v.get('roll_no',''),   v.get('level',''),
                v.get('course',''),    v.get('year',''),
                v.get('jc_year',''),   v.get('jc_stream',''),
                v.get('purpose',''),   str(v.get('entry_time','')),
                str(v.get('exit_time','')), str(v.get('visit_date','')),
                v.get('visit_day',''), duration
            ]

        filename_base = f"library_{label}"

        # ── CSV ────────────────────────────────────────────────────
        if format_type == 'csv':
            output = io.StringIO()
            w = csv.writer(output)

            # meta header
            w.writerow([f"Library Visitor Report — {label}"])
            w.writerow([f"Exported: {datetime.now().strftime('%d %b %Y %H:%M')}",
                        f"Total records: {len(visitors)}"])
            w.writerow([])
            w.writerow(HEADERS)
            for v in visitors:
                w.writerow(build_row(v))

            output.seek(0)
            return send_file(
                io.BytesIO(output.getvalue().encode('utf-8')),
                mimetype='text/csv',
                as_attachment=True,
                download_name=f"{filename_base}.csv"
            )

        # ── EXCEL ──────────────────────────────────────────────────
        elif format_type == 'excel':
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()

            # ── helper: style a sheet ──────────────────────────────
            def style_sheet(ws, rows_data, sheet_label):
                # Title row
                ws.append([f"Library Visitor Report — {sheet_label}"])
                ws.append([f"Exported: {datetime.now().strftime('%d %b %Y %H:%M')}",
                           f"Total records: {len(rows_data)}"])
                ws.append([])  # blank row
                ws.append(HEADERS)

                header_row = 4
                title_fill  = PatternFill("solid", fgColor="4F46E5")
                header_fill = PatternFill("solid", fgColor="6366F1")
                alt_fill    = PatternFill("solid", fgColor="EEF2FF")
                white_fill  = PatternFill("solid", fgColor="FFFFFF")
                thin = Side(style='thin', color='C7D2FE')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # Style title
                title_cell = ws.cell(row=1, column=1)
                title_cell.font = Font(bold=True, size=13, color="FFFFFF")
                title_cell.fill = title_fill
                title_cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[1].height = 24
                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=len(HEADERS))

                # Style subtitle
                ws.cell(row=2, column=1).font = Font(italic=True, color="64748B", size=10)

                # Style header row
                for col, h in enumerate(HEADERS, 1):
                    cell = ws.cell(row=header_row, column=col)
                    cell.value = h
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                ws.row_dimensions[header_row].height = 22

                # Data rows
                for i, row in enumerate(rows_data):
                    r = header_row + 1 + i
                    fill = alt_fill if i % 2 == 0 else white_fill
                    for col, val in enumerate(row, 1):
                        cell = ws.cell(row=r, column=col, value=val)
                        cell.fill = fill
                        cell.border = border
                        cell.alignment = Alignment(vertical='center')
                        cell.font = Font(size=10)

                # Auto-fit column widths
                col_widths = [max(len(str(h)), 8) for h in HEADERS]
                for row in rows_data:
                    for i, val in enumerate(row):
                        col_widths[i] = min(max(col_widths[i], len(str(val))), 30)
                for i, w in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w + 2

                # Freeze panes below header
                ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

                # Auto-filter on header row
                ws.auto_filter.ref = (
                    f"A{header_row}:{get_column_letter(len(HEADERS))}{header_row + len(rows_data)}"
                )

            # ── decide sheet layout ────────────────────────────────
            all_rows = [build_row(v) for v in visitors]

            if scope == 'all':
                # Sheet 1: All data
                ws_all = wb.active
                ws_all.title = "All Time"
                style_sheet(ws_all, all_rows, "All Time")

                # Sheet 2: Monthly summary
                from collections import defaultdict
                monthly = defaultdict(list)
                for v in visitors:
                    d = str(v.get('visit_date', ''))
                    if len(d) >= 7:
                        monthly[d[:7]].append(v)

                ws_sum = wb.create_sheet("Monthly Summary")
                ws_sum.append(["Month", "Total Visitors", "JC", "UG", "PG",
                               "Avg Duration (min)"])
                hfill = PatternFill("solid", fgColor="6366F1")
                for col in range(1, 7):
                    c = ws_sum.cell(row=1, column=col)
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = hfill
                for m_key in sorted(monthly.keys(), reverse=True):
                    mvs = monthly[m_key]
                    jc = sum(1 for v in mvs if v.get('level') == 'JC')
                    ug = sum(1 for v in mvs if v.get('level') == 'UG')
                    pg = sum(1 for v in mvs if v.get('level') == 'PG')
                    durs = []
                    for v in mvs:
                        if v.get('entry_time') and v.get('exit_time'):
                            try:
                                from datetime import datetime as dt
                                e = dt.strptime(str(v['entry_time']), '%H:%M:%S')
                                x = dt.strptime(str(v['exit_time']),  '%H:%M:%S')
                                durs.append((x - e).total_seconds() / 60)
                            except Exception:
                                pass
                    avg = round(sum(durs) / len(durs), 1) if durs else ''
                    ws_sum.append([m_key, len(mvs), jc, ug, pg, avg])

                ws_sum.column_dimensions['A'].width = 14
                for col in 'BCDEF':
                    ws_sum.column_dimensions[col].width = 18

            elif scope == 'year' and year:
                # One sheet per month in that year
                from collections import defaultdict
                monthly = defaultdict(list)
                for v in visitors:
                    d = str(v.get('visit_date', ''))
                    if len(d) >= 7:
                        monthly[d[:7]].append(v)

                first = True
                for m_key in sorted(monthly.keys()):
                    rows = [build_row(v) for v in monthly[m_key]]
                    import calendar
                    m_name = calendar.month_name[int(m_key.split('-')[1])]
                    if first:
                        ws = wb.active
                        ws.title = m_name
                        first = False
                    else:
                        ws = wb.create_sheet(m_name)
                    style_sheet(ws, rows, f"{m_name} {year}")

                if first:  # no data
                    ws = wb.active
                    ws.title = year
                    style_sheet(ws, [], year)

            else:
                # Single sheet (month / custom range)
                ws = wb.active
                ws.title = label[:31]  # Excel sheet name limit
                style_sheet(ws, all_rows, label)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f"{filename_base}.xlsx"
            )

        else:
            return jsonify({"error": "Invalid format"}), 400

    except Exception as e:
        print(f"Export error: {e}")
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Export failed: {str(e)}"}), 500

@admin_bp.route('/bulk_actions', methods=['POST'])
@login_required
def bulk_actions():
    """Perform bulk actions on visitors"""
    try:
        data = request.json
        action = data.get('action')
        visitor_ids = data.get('visitor_ids', [])
        
        if not visitor_ids:
            return jsonify({"error": "No visitors selected"}), 400
        
        if action == 'mark_exit':
            success_count = 0
            for visitor_id in visitor_ids:
                try:
                    # ✅ USE DIRECT API
                    result = Database.update_exit_by_id(visitor_id)
                    if result:
                        success_count += 1
                except:
                    pass
            
            return jsonify({
                "success": True,
                "message": f"Marked {success_count} visitors as exited"
            }), 200
        
        elif action == 'delete':
            success_count = 0
            for visitor_id in visitor_ids:
                try:
                    # ✅ USE DIRECT API
                    result = Database.delete_visitor(visitor_id)
                    if result:
                        success_count += 1
                except:
                    pass
            
            return jsonify({
                "success": True,
                "message": f"Deleted {success_count} visitors"
            }), 200
        
        else:
            return jsonify({"error": "Invalid action"}), 400
            
    except Exception as e:
        print(f"Bulk action error: {e}")
        return jsonify({"error": "Bulk action failed"}), 500


